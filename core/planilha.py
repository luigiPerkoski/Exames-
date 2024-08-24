from openpyxl import load_workbook


def ler_com_openpyxl(caminho_arquivo, nome_planilha):

    wb = load_workbook(caminho_arquivo, data_only=True)
    sheet = wb[nome_planilha]

    dados = []
    for row in sheet.iter_rows(min_row=2, values_only=False):
        apelido = row[0].value
        nome = row[1].value

        # Resolver a fórmula em 'price'
        cell_price = row[4]
        if cell_price.data_type == 'f':  # Se a célula contém uma fórmula
            price = cell_price.internal_value  # Obtenha o resultado da fórmula
        else:
            price = cell_price.value  # Caso contrário, pegue o valor direto

        # Formatar 'price' para duas casas decimais
        price = f'{price:.2f}' if isinstance(price, (int, float)) else price

        linha = {'surname': apelido, 'name': nome, 'price': price}
        dados.append(linha)


    return dados

def main():
    caminho_arquivo = r"C:\Luigi\Document\Code\Python\Project\Exames\TabelaLab.xlsx" 
    nome_planilha = 'Planilha1'

    dados = ler_com_openpyxl(caminho_arquivo, nome_planilha)
    print(dados[2])

if __name__ == "__main__":
    main()
